import openpyxl


def reader(f_name: str = 'data.xlsx') -> openpyxl.load_workbook:
    wb = openpyxl.load_workbook(f_name)
    return wb


def analyse_cities_salaries(wb: openpyxl.load_workbook) -> int:
    sheet = wb.active
    rows = sheet.max_row
    columns = sheet.max_column

    print(rows, columns)
    cities = {}
    for i in range(2, rows + 1):
        print((q := sheet.cell(row=i, column=1).value))
        cities[q] = cities.setdefault(q, [])
        for j in range(2, columns+1):
            cities[q].append(sheet.cell(row=i, column=j).value)
        cities[q].sort()
    print(cities)
    print(max(cities.items(),key=lambda x: x[1][3]))
    return max(cities.values(),key=lambda x: x[3])[3]


def analyse_prof_salaries(wb: openpyxl.load_workbook) -> int:
    sheet = wb.active
    rows = sheet.max_row
    columns = sheet.max_column

    print(rows, columns)
    professions = {}
    for j in range(2, columns+1):
        print((q := sheet.cell(row=1,column=j).value))
        professions[q] = professions.setdefault(q,[])
        for i in range(2, rows):
            professions[q].append(sheet.cell(row=i, column=j).value)
    print(professions)
    print(max(professions.items(), key= lambda x: sum(x[1])/len(x[1])))


def main():
    data = reader()
    analyse_prof_salaries(data)
    analyse_cities_salaries(data)


if __name__ == '__main__':
    main()